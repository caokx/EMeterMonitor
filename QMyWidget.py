import datetime
import os
import sys
import time

import openpyxl
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtCore import QTimer

from FCS import FCS
from PlotWindow import PlotWindow
from PlotWindowConjunction import PlotWindowConjunction
from QGetDataThread import QGetDataThread
from QPort import QPort
from TimedSaveThread import TimedSaveThread
from Ui_MainWindow import Ui_MainWindow
from WriteExcelThread import WriteExcelThread
from plotHistory2 import PlotHistoryWindowConjunction
from PlotHistoryWindow import PlotHistoryWindow

DATADIC = {}
DATADICKEY = []

DATADIC_HISTORY = {}
DATADICKEY_HISTORY = []


class QMyWidget(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        '''信号量'''
        self.serialIsOpen = -1
        self.serialIsCommunicated = -1
        self.serialIsCommunicating = -1

        self.ui = Ui_MainWindow()  # 创建UI类
        self.ui.setupUi(self)  # 添加UI的界面
        self.ui.setUi()  # 设置UI界面

        self.port = QPort()
        self.serial = self.port.serial

        self.binding()  # 绑定按钮与槽函数

        '''端口号扫描'''
        portNames = QPort.getPortNames()
        self.ui.tab0Con0H0Com.clear()
        self.ui.tab0Con0H0Com.addItems(portNames)
        self.scanPortsTimer()  # 开启扫描

    def binding(self):
        '''tab0绑定'''
        self.ui.tab0Con1H0Bn.clicked.connect(self.openSerial)
        self.ui.tab0Con1H2Bn.clicked.connect(self.testCommunication)
        self.ui.tab0Con1H3Bn.clicked.connect(self.communicate)
        self.ui.tab0Con1H1Bn.clicked.connect(self.closeSerial)
        self.ui.tab0Con1H4Bn.clicked.connect(self.closeSerial)
        '''tab1绑定'''
        self.ui.tab1Con0Bn2.clicked.connect(self.voltagePlotWindow)
        self.ui.tab1Con1Bn2.clicked.connect(self.currentPlotWindow)
        self.ui.tab1Con2Bn2.clicked.connect(self.powerPlotWindow)
        self.ui.tab1Con3Bn2.clicked.connect(self.powerFactorPlotWindow)
        self.ui.tab1Con4Bn2.clicked.connect(self.energyPlotWindow)
        '''tab2绑定'''
        self.ui.tab2Con_1Bn0.clicked.connect(self.beforePlotHistory)
        self.ui.tab2Con0Bn2.clicked.connect(self.plotHistoryVoltage)
        self.ui.tab2Con1Bn2.clicked.connect(self.plotHistoryCurrent)
        self.ui.tab2Con2Bn2.clicked.connect(self.plotHistoryPower)
        self.ui.tab2Con3Bn2.clicked.connect(self.plotHistoryPowerFactor)
        self.ui.tab2Con4Bn2.clicked.connect(self.plotHistoryEnergy)


    def plotHistoryEnergy(self):
        self.plotE = PlotHistoryWindow(title='电能量' + self.ui.tab2Con4Com1.currentText(),
                                       chose=self.ui.tab2Con4Com1.currentIndex() + 14,
                                       beginTime=self.ui.tab2Con_1Lab3.text(),
                                       endTime=self.ui.tab2Con_1Lab5.text(),
                                       DATADIC=DATADIC_HISTORY, DATADICKEY=DATADICKEY_HISTORY)
        self.plotE.show()
    def plotHistoryPowerFactor(self):
        if self.ui.tab2Con3Com1.currentIndex() < 4:
            self.plotPF = PlotHistoryWindow(title='功率因数' + self.ui.tab2Con3Com1.currentText(),
                                           chose=self.ui.tab2Con3Com1.currentIndex()+10,
                                           beginTime=self.ui.tab2Con_1Lab3.text(),
                                           endTime=self.ui.tab2Con_1Lab5.text(),
                                           DATADIC=DATADIC_HISTORY, DATADICKEY=DATADICKEY_HISTORY)
            self.plotPF.show()
        else:
            self.plotPF = PlotHistoryWindowConjunction(title='功率因数三相', chose=[11, 12, 13],
                                                      beginTime=self.ui.tab2Con_1Lab3.text(),
                                                      endTime=self.ui.tab2Con_1Lab5.text(),
                                                      DATADIC=DATADIC_HISTORY, DATADICKEY=DATADICKEY_HISTORY)

            self.plotPF.show()
    def plotHistoryPower(self):
        if self.ui.tab2Con2Com1.currentIndex() < 4:
            self.plotP = PlotHistoryWindow(title='功率' + self.ui.tab2Con2Com1.currentText(),
                                           chose=self.ui.tab2Con2Com1.currentIndex()+6,
                                           beginTime=self.ui.tab2Con_1Lab3.text(),
                                           endTime=self.ui.tab2Con_1Lab5.text(),
                                           DATADIC=DATADIC_HISTORY, DATADICKEY=DATADICKEY_HISTORY)
            self.plotP.show()
        else:
            self.plotP = PlotHistoryWindowConjunction(title='功率三相', chose=[7, 8, 9],
                                                      beginTime=self.ui.tab2Con_1Lab3.text(),
                                                      endTime=self.ui.tab2Con_1Lab5.text(),
                                                      DATADIC=DATADIC_HISTORY, DATADICKEY=DATADICKEY_HISTORY)

            self.plotP.show()

    def plotHistoryCurrent(self):
        if self.ui.tab2Con1Com1.currentIndex() < 3:
            self.plotC = PlotHistoryWindow(title='电流' + self.ui.tab2Con1Com1.currentText(),
                                           chose=self.ui.tab2Con1Com1.currentIndex()+3,
                                           beginTime=self.ui.tab2Con_1Lab3.text(),
                                           endTime=self.ui.tab2Con_1Lab5.text(),
                                           DATADIC=DATADIC_HISTORY, DATADICKEY=DATADICKEY_HISTORY)
            self.plotC.show()
        else:
            self.plotC = PlotHistoryWindowConjunction(title='电流三相', chose=[3, 4, 5],
                                                      beginTime=self.ui.tab2Con_1Lab3.text(),
                                                      endTime=self.ui.tab2Con_1Lab5.text(),
                                                      DATADIC=DATADIC_HISTORY, DATADICKEY=DATADICKEY_HISTORY)

            self.plotC.show()

    def plotHistoryVoltage(self):
        if self.ui.tab2Con0Com1.currentIndex() < 3:
            self.plotV = PlotHistoryWindow(title='电压'+self.ui.tab2Con0Com1.currentText(),
                                           chose=self.ui.tab2Con0Com1.currentIndex(),
                                           beginTime=self.ui.tab2Con_1Lab3.text(),
                                           endTime=self.ui.tab2Con_1Lab5.text(),
                                           DATADIC=DATADIC_HISTORY, DATADICKEY=DATADICKEY_HISTORY)
            self.plotV.show()
        else:
            self.plotV = PlotHistoryWindowConjunction(title='电压三相', chose=[0, 1, 2],
                                           beginTime=self.ui.tab2Con_1Lab3.text(),
                                           endTime=self.ui.tab2Con_1Lab5.text(),
                                           DATADIC=DATADIC_HISTORY, DATADICKEY=DATADICKEY_HISTORY)

            self.plotV.show()

    def beforePlotHistory(self):
        DATADIC_HISTORY.clear()
        DATADICKEY_HISTORY.clear()
        current_directory = os.getcwd()
        history_directory = current_directory + '\history'
        files = [f for f in os.listdir(history_directory) if
                 os.path.isfile(os.path.join(history_directory, f))]
        """遍历文件加载数据"""
        for file in files:
            path = history_directory + "\\" + str(file)
            """如果文件可以打开"""
            if os.access(path=path, mode=os.F_OK):
                xlsx = openpyxl.load_workbook(path)
                sheet = xlsx.active
                if sheet.max_row < 2 or sheet.max_column != 16:
                    xlsx.close()
                    continue
                """加载数据"""
                for row in sheet.iter_rows(min_row=2, values_only=True):  # 仅读取单元格的值
                    DATADICKEY_HISTORY.append(row[0])
                    DATADIC_HISTORY[DATADICKEY_HISTORY[-1]] = row[1:]
                xlsx.close()
        """时间排序"""
        # DATADICKEY_HISTORY.sort()
        self.ui.tab2Con_1Lab3.setText(str(DATADICKEY_HISTORY[0]))
        self.ui.tab2Con_1Lab5.setText(str(DATADICKEY_HISTORY[-1]))

    def energyPlotWindow(self):
        self.pe = PlotWindow(title=self.ui.tab1Con4Com1.currentText(),
                             scope=int(self.ui.tab1Con4Com3.currentText()),
                             chose=self.ui.tab1Con4Com1.currentIndex() + 14,
                             DATADIC=DATADIC, DATADICKEY=DATADICKEY)
        self.pe.show()

    def powerFactorPlotWindow(self):
        if self.ui.tab1Con3Com1.currentIndex() < 4:
            self.ppf0 = PlotWindow(title='功率因数' + self.ui.tab1Con3Com1.currentText(),
                                   scope=int(self.ui.tab1Con3Com3.currentText()),
                                   chose=self.ui.tab1Con3Com1.currentIndex() + 10,
                                   DATADIC=DATADIC, DATADICKEY=DATADICKEY)

            self.ppf0.show()
        else:
            self.ppf3 = PlotWindowConjunction(title='功率因数三相', scope=int(self.ui.tab1Con3Com3.currentText()),
                                              chose=[11, 12, 13],
                                              DATADIC=DATADIC, DATADICKEY=DATADICKEY)
            self.ppf3.show()

    def powerPlotWindow(self):
        if self.ui.tab1Con2Com1.currentIndex() < 4:
            self.pp0 = PlotWindow(title='功率' + self.ui.tab1Con2Com1.currentText(),
                                  scope=int(self.ui.tab1Con2Com3.currentText()),
                                  chose=self.ui.tab1Con2Com1.currentIndex() + 6,
                                  DATADIC=DATADIC, DATADICKEY=DATADICKEY)

            self.pp0.show()
        else:
            self.pp3 = PlotWindowConjunction(title='功率三相', scope=int(self.ui.tab1Con2Com3.currentText()),
                                             chose=[7, 8, 9],
                                             DATADIC=DATADIC, DATADICKEY=DATADICKEY)
            self.pp3.show()

    def currentPlotWindow(self):
        if self.ui.tab1Con1Com1.currentIndex() < 3:
            self.pc0 = PlotWindow(title='电流' + self.ui.tab1Con1Com1.currentText(),
                                  scope=int(self.ui.tab1Con1Com3.currentText()),
                                  chose=self.ui.tab1Con1Com1.currentIndex() + 3,
                                  DATADIC=DATADIC, DATADICKEY=DATADICKEY)

            self.pc0.show()
        else:
            self.pc3 = PlotWindowConjunction(title='电流三相', scope=int(self.ui.tab1Con1Com3.currentText()),
                                             chose=[3, 4, 5],
                                             DATADIC=DATADIC, DATADICKEY=DATADICKEY)
            self.pc3.show()

    def voltagePlotWindow(self):
        if self.ui.tab1Con0Com1.currentIndex() < 3:
            self.pv0 = PlotWindow(title='电压' + self.ui.tab1Con0Com1.currentText(),
                                  scope=int(self.ui.tab1Con0Com3.currentText()),
                                  chose=self.ui.tab1Con0Com1.currentIndex(),
                                  DATADIC=DATADIC, DATADICKEY=DATADICKEY)

            self.pv0.show()
        else:
            self.pv3 = PlotWindowConjunction(title='电压三相', scope=int(self.ui.tab1Con0Com3.currentText()),
                                             chose=[0, 1, 2],
                                             DATADIC=DATADIC, DATADICKEY=DATADICKEY)
            self.pv3.show()

    def scanPortsTimer(self):
        # 端口检测线程
        self.scanTimer = QTimer()
        self.scanTimer.timeout.connect(self.getPortNames)
        self.scanTimer.start(500)

    def getPortNames(self):
        portNames = QPort.getPortNames()
        if self.serialIsOpen == 1 and self.serial.port not in portNames:
            self.closeSerial()
            self.serialIsOpen = -1
            self.serialIsCommunicating = -1

        if self.serialIsOpen == -1:
            # 若串口拔出，则在combox删除下拉栏
            for index in range(self.ui.tab0Con0H0Com.count()):
                if self.ui.tab0Con0H0Com.itemText(index) not in portNames:
                    self.ui.tab0Con0H0Com.removeItem(index)
            # 若有新的串口，则在下拉栏添加
            if self.ui.tab0Con0H0Com.count() < len(portNames):
                for index in range(self.ui.tab0Con0H0Com.count()):
                    portNames = [x for x in portNames if x != self.ui.tab0Con0H0Com.itemText(index)]
                for item in portNames:
                    self.ui.tab0Con0H0Com.addItem(item)

    # 重新打开串口
    def openSerial(self):
        port = self.ui.tab0Con0H0Com.currentText()
        baudrate = int(self.ui.tab0Con0H1Com.currentText())
        parity = self.ui.tab0Con0H2Com.currentText()
        bytesize = int(self.ui.tab0Con0H3Com.currentText())
        stopbits = int(self.ui.tab0Con0H4Com.currentText())

        self.port.setSerial(port=port, baudrate=baudrate, parity=parity[0],
                            stopbits=stopbits, bytesize=bytesize)

        try:
            self.serial.open()
        except IOError:
            sys.stdout.write('串口被占用')
            return

        self.serialIsOpen = 1
        sys.stdout.write('串口打开中')
        self.ui.tab0Con1H0Lab.setText('开放')
        self.ui.tab0Con1H0Lab.setStyleSheet('color:green')
        self.ui.tab0Con0H0Com.setEnabled(False)
        self.ui.tab0Con0H1Com.setEnabled(False)
        self.ui.tab0Con0H2Com.setEnabled(False)
        self.ui.tab0Con0H3Com.setEnabled(False)
        self.ui.tab0Con0H4Com.setEnabled(False)
        self.ui.tab0Con1H0Bn.setEnabled(False)
        self.ui.tab0Con1H1Bn.setEnabled(True)
        self.ui.tab0Con1H2Bn.setEnabled(True)

    def closeSerial(self):
        self.serialIsOpen = -1
        self.serialIsCommunicated = -1
        self.ui.tab0Con0H0Com.setEnabled(True)
        self.ui.tab0Con0H1Com.setEnabled(True)
        self.ui.tab0Con0H2Com.setEnabled(True)
        self.ui.tab0Con0H3Com.setEnabled(True)
        self.ui.tab0Con0H4Com.setEnabled(True)
        self.ui.tab0Con1H0Bn.setEnabled(True)
        self.ui.tab0Con1H1Bn.setEnabled(True)
        self.ui.tab0Con1H2Bn.setEnabled(False)
        self.ui.tab0Con1H0Lab.setText('关闭')
        self.ui.tab0Con1H0Lab.setStyleSheet('color:black')
        if self.serialIsCommunicating == 1:
            try:
                self.thread.terminate()  # 停止采集数据
                self.ui.tab0Con1H3Lab.setText('停止')
                self.ui.tab0Con1H3Lab.setStyleSheet('color:black')
                self.ui.tab0Con1H4Bn.setEnabled(False)
                self.ui.tab0Con1H2Lab.setText('')
                while True:
                    if self.writeExcelThread.excelIsWriting == 1:
                        pass
                    if self.writeExcelThread.excelIsWriting == -1:
                        self.writeExcelThread.terminate()
                        break
            except IOError:
                sys.stdout.write('进程终止出错')
        self.serial.close()
        sys.stdout.write('串口关闭')

    def testCommunication(self):
        mutex = QtCore.QMutex()
        mutex.lock()
        if self.serialIsOpen == 1:
            message = '68 17 00 43 45 AA AA AA AA AA AA a1 D8 FB 05 01 02 40 01 02 00 00 BB 0B 16'
            # message = '68 17 00 43 45 AA AA AA AA AA AA a1 D8 FB 05 01 02 00 10 02 01 00 5b 0c 16'
            sys.stdout.write('发送报文->' + message)
            self.writeMessage = bytes.fromhex(message)
            self.serial.write(self.writeMessage)
            time.sleep(0.4)
            lenOfRead = self.serial.inWaiting()  # 获取缓冲数据（接收数据）长度
            if lenOfRead:
                self.readMessage = self.serial.read(lenOfRead)
                if FCS.testFcs(cp=self.readMessage[5:-1]):
                    read = ''.join(map(lambda x: (" " if len(hex(x)) >= 4 else " 0") + hex(x)[2:],
                                       self.readMessage))
                    sys.stdout.write('接收报文<-' + read)
                    self.ui.tab0Con1H2Lab.setText('正常')
                    self.ui.tab0Con1H2Lab.setStyleSheet('color:green')
                    self.address = self.readMessage[28:34]
                    self.serialIsCommunicated = 1
                    self.ui.tab0Con1H3Bn.setEnabled(True)
            else:
                sys.stdout.write('没有收到电表返回报文')
                self.ui.tab0Con1H0Lab.setText('通信异常')
                self.serialIsCommunicated = -1

        else:
            self.ui.tab0Con1H0Lab.setText('关闭')
            sys.stdout.write('请先打开串口')
        mutex.unlock()

    def communicate(self):
        """更换数据采集类型时，只需要修改getData.message、selectData和writeExcelThread即可"""
        # message = '68 24 00 43 45 AA AA AA AA AA AA A1 A4 96 05 02 00 04 20 ' \
        #           '00 02 00 20 01 02 00 20 04 02 00 20 05 02 00 00 31 45 16'
        message = '68 28 00 43 45 aa aa aa aa aa aa a1 5b 74 05 02 00 05 20 00' \
                  ' 02 00 20 01 02 00 20 04 02 00 20 0a  02 00 00 10 02 01 00 72 e4 16'
        if self.serialIsCommunicated == 1:
            self.thread = QGetDataThread(self.serial, message)
            self.thread.signal_.connect(self.pickData)
            self.thread.start()
            sys.stdout.write('数据采集中...')
            self.ui.tab0Con1H3Lab.setText('采集')
            self.ui.tab0Con1H3Lab.setStyleSheet('color:green')
            self.ui.tab0Con1H1Bn.setEnabled(False)
            self.ui.tab0Con1H2Bn.setEnabled(False)
            self.ui.tab0Con1H3Bn.setEnabled(False)
            self.ui.tab0Con1H4Bn.setEnabled(True)
            self.writeExcelThread = WriteExcelThread(DATADIC=DATADIC)
            self.writeExcelThread.start()
            self.serialIsCommunicating = 1

        else:
            sys.stdout.write('请先测试能否正常通信')

    def pickData(self, read):
        current_ = datetime.datetime.now()
        currentTime = current_.strftime('%Y-%m-%d %H:%M:%S')
        # 如果当前时间是00：00：00，就自动保存数据一次数据在history文件夹，并清空DIC
        if (current_.hour == 13 and current_.minute == 0 and current_.second == 0) or (
                current_.hour == 1 and current_.minute == 0 and current_.second == 0):
            self.timedSave = TimedSaveThread(DATADIC=DATADIC.copy(), current_=current_)
            self.timedSave.start()
            DATADIC.clear()
            DATADICKEY.clear()
        voltageA = int(read[30:32].hex(), 16) / 10
        voltageB = int(read[33:35].hex(), 16) / 10
        voltageC = int(read[36:38].hex(), 16) / 10
        currentA = int(read[46:50].hex(), 16) / 1000
        currentB = int(read[51:55].hex(), 16) / 1000
        currentC = int(read[56:60].hex(), 16) / 1000
        powerTotal = int(read[68:72].hex(), 16) / 10
        powerA = int(read[73:77].hex(), 16) / 10
        powerB = int(read[78:82].hex(), 16) / 10
        powerC = int(read[83:87].hex(), 16) / 10
        powerFactorTotal = int(read[95:97].hex(), 16) / 1000
        powerFactorA = int(read[98:100].hex(), 16) / 1000
        powerFactorB = int(read[101:103].hex(), 16) / 1000
        powerFactorC = int(read[104:106].hex(), 16) / 1000
        energy = int(read[112:116].hex(), 16) / 100
        sys.stdout.write('电压:' + str(voltageA) + "| " + str(voltageB) + "| " + str(voltageC) + '\n'
                         + '电流:' + str(currentA) + '| ' + str(currentB) + '| ' + str(currentC) + '\n'
                         + '功率:' + str(powerTotal) + '| ' + str(powerA) + '| ' + str(powerB) + '| ' + str(
            powerC) + '\n'
                         + '功率因数:' + str(powerFactorTotal) + '| ' + str(powerFactorA) + '| ' + str(
            powerFactorB) + '| ' + str(powerFactorC) + '\n'
                         + '电能量：' + str(energy)
                         )
        DATADIC[currentTime] = [voltageA, voltageB, voltageC,
                                currentA, currentB, currentC,
                                powerTotal, powerA, powerB, powerC,
                                powerFactorTotal, powerFactorA, powerFactorB, powerFactorC,
                                energy]
        DATADICKEY.append(currentTime)
