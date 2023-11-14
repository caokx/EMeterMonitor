import os
import sys
import time

import openpyxl
from PyQt5 import QtCore




class WriteExcelThread(QtCore.QThread):
    def __init__(self, DATADIC={}):
        super().__init__()
        self.excelIsWriting = -1
        self.DATADIC = DATADIC

    def run(self):
        while True:
            time.sleep(30)
            self.write()

    def write(self):
        data = self.DATADIC.copy()
        try:
            '''
                os.F_OK: 检查文件是否存在;
                os.R_OK: 检查文件是否可读;
                os.W_OK: 检查文件是否可以写入;
                os.X_OK: 检查文件是否可以执行
            '''
            if os.access(path='tmp/data.xlsx', mode=os.F_OK):
                if os.access(path='tmp/data.xlsx', mode=os.W_OK):
                    # sys.stdout.write("文件存在且可写")
                    # 读取
                    xlsx = openpyxl.load_workbook("tmp/data.xlsx")
                    self.excelIsWriting = 1
                    table = xlsx.active
                    table.delete_rows(2, table.max_row)
                    # 写入
                    i = 2
                    for key in data.keys():
                        table['A' + str(i)] = key
                        table['B' + str(i)] = data[key][0]
                        table['C' + str(i)] = data[key][1]
                        table['D' + str(i)] = data[key][2]
                        table['E' + str(i)] = data[key][3]
                        table['F' + str(i)] = data[key][4]
                        table['G' + str(i)] = data[key][5]
                        table['H' + str(i)] = data[key][6]
                        table['I' + str(i)] = data[key][7]
                        table['J' + str(i)] = data[key][8]
                        table['K' + str(i)] = data[key][9]
                        table['L' + str(i)] = data[key][10]
                        table['M' + str(i)] = data[key][11]
                        table['N' + str(i)] = data[key][12]
                        table['O' + str(i)] = data[key][13]
                        table['P' + str(i)] = data[key][14]
                        i += 1
                    # 保存
                    xlsx.save("tmp/data.xlsx")
                    xlsx.close()
                    self.excelIsWriting = -1

                else:   # 文件存在但是不可读
                    sys.stdout.write("文件存在但不可写")


            else:
                sys.stdout.write("创建文件")
                xlsx = openpyxl.Workbook()
                self.excelIsWriting = 1
                # 获取活跃的工作表
                table = xlsx.active
                table.title = "record"

                # 表头
                table['A1'] = "时间"
                table['B1'] = "电压A相(V)"
                table['C1'] = "电压B相(V)"
                table['D1'] = "电压C相(V)"
                table['E1'] = "电流A相(A)"
                table['F1'] = "电流B相(A)"
                table['G1'] = "电流C相(A)"
                table['H1'] = "总功率(W)"
                table['I1'] = "功率A相(W)"
                table['J1'] = "功率B相(W)"
                table['K1'] = "功率C相(W)"
                table['L1'] = "总功率因数"
                table['M1'] = "功率因数A相"
                table['N1'] = "功率因数B相"
                table['O1'] = "功率因数C相"
                table['P1'] = "电能量(kw*h)"




                # 写入
                i = 2
                for key in data.keys():
                    table['A' + str(i)] = key
                    table['B' + str(i)] = data[key][0]
                    table['C' + str(i)] = data[key][1]
                    table['D' + str(i)] = data[key][2]
                    table['E' + str(i)] = data[key][3]
                    table['F' + str(i)] = data[key][4]
                    table['G' + str(i)] = data[key][5]
                    table['H' + str(i)] = data[key][6]
                    table['I' + str(i)] = data[key][7]
                    table['J' + str(i)] = data[key][8]
                    table['K' + str(i)] = data[key][9]
                    table['L' + str(i)] = data[key][10]
                    table['M' + str(i)] = data[key][11]
                    table['N' + str(i)] = data[key][12]
                    table['O' + str(i)] = data[key][13]
                    table['P' + str(i)] = data[key][14]
                    i += 1
                # 保存
                xlsx.save("tmp/data.xlsx")
                xlsx.close()
                self.excelIsWriting = -1


        except:
            sys.stdout.write("文件操作错误")
            pass
